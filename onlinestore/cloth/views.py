from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt  # Добавлен этот импорт
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Count, Avg, Min, Max, Sum
from django.contrib import messages
from django.core.paginator import Paginator
from django.utils import timezone
from django.conf import settings  # Добавлен этот импорт
from datetime import timedelta
from .models import (
    Product, ProductVariant, Category, Cart, CartItem,
    Wishlist, Order, OrderStatus, Review, User, Role,
    DeliveryMethod, Transaction, TransactionStatus,
    Size, Color, Gender, ProductImage, OrderItem
)
from .forms import RegisterForm, LoginForm, CheckoutForm, ReviewForm, UserProfileForm, ChangePasswordForm
from .payments import PaymentService  # Импорт сервиса платежей
import logging
from decimal import Decimal
import uuid
import csv
import json  # Добавлен этот импорт
from .utils import send_verification_email, send_password_reset_email
from .models import EmailVerification

logger = logging.getLogger(__name__)

# Создаем экземпляр сервиса платежей
payment_service = PaymentService()


# Проверка на модератора или админа
def is_moderator(user):
    return user.is_authenticated and (user.role.name in ['moderator', 'admin'] or user.is_staff)


# Проверка на администратора
def is_admin(user):
    return user.is_authenticated and (user.role.name == 'admin' or user.is_superuser)


# --- ГЛАВНАЯ И КАТАЛОГ ---
def home(request):
    """Главная страница с новинками"""
    products = Product.objects.filter(is_active=True).select_related(
        'category'
    ).prefetch_related(
        'variants', 'images'
    ).order_by('-created_at')[:8]

    new_products = Product.objects.filter(
        is_active=True, is_new=True
    ).select_related('category')[:4]

    bestsellers = Product.objects.filter(
        is_active=True, is_bestseller=True
    ).select_related('category')[:4]

    # Получаем список ID товаров в избранном для текущего пользователя
    wishlist_ids = []
    if request.user.is_authenticated:
        wishlist_ids = list(request.user.wishlist.values_list('product_id', flat=True))

    return render(request, "pages/home.html", {
        "products": products,
        "new_products": new_products,
        "bestsellers": bestsellers,
        "wishlist_ids": wishlist_ids,  # Добавляем в контекст
    })


def catalog(request):
    """Каталог товаров с фильтрацией"""
    categories = Category.objects.filter(is_active=True)
    products = Product.objects.filter(is_active=True).select_related(
        'category', 'gender'
    ).prefetch_related('variants', 'images')

    # Получаем все размеры и цвета для фильтров
    sizes = Size.objects.all().order_by('order')
    colors = Color.objects.all()

    # Получаем выбранные значения из GET параметров
    selected_sizes = request.GET.getlist('size')
    selected_colors = request.GET.getlist('color')

    # Фильтрация по категории
    cat_slug = request.GET.get('category')
    if cat_slug:
        products = products.filter(category__slug=cat_slug)

    # Фильтрация по размеру
    if selected_sizes:
        products = products.filter(variants__size__name__in=selected_sizes)

    # Фильтрация по цвету
    if selected_colors:
        products = products.filter(variants__color__name__in=selected_colors)

    # Фильтрация по цене
    min_price = request.GET.get('min_price')
    if min_price:
        products = products.filter(price__gte=min_price)

    max_price = request.GET.get('max_price')
    if max_price:
        products = products.filter(price__lte=max_price)

    # Поиск
    q = request.GET.get('q')
    if q:
        products = products.filter(
            Q(name__icontains=q) |
            Q(description__icontains=q) |
            Q(category__name__icontains=q)
        )

    # Сортировка
    sort = request.GET.get('sort', '-created_at')
    if sort in ['price', '-price', 'name', '-name', 'created_at', '-created_at']:
        products = products.order_by(sort)

    # Пагинация
    paginator = Paginator(products.distinct(), 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    # Статистика цен
    price_stats = Product.objects.filter(is_active=True).aggregate(
        min_price=Min('price'),
        max_price=Max('price')
    )

    # Получаем список ID товаров в избранном для текущего пользователя
    wishlist_ids = []
    if request.user.is_authenticated:
        wishlist_ids = list(request.user.wishlist.values_list('product_id', flat=True))

    context = {
        'page_obj': page_obj,
        'products': page_obj,
        'categories': categories,
        'sizes': sizes,
        'colors': colors,
        'selected_sizes': selected_sizes,
        'selected_colors': selected_colors,
        'min_price_global': price_stats['min_price'] or 0,
        'max_price_global': price_stats['max_price'] or 100000,
        'selected_category': cat_slug,
        'wishlist_ids': wishlist_ids,  # Добавляем в контекст
    }

    return render(request, "pages/catalog.html", context)

def product_detail(request, slug):
    product = get_object_or_404(
        Product.objects.select_related('category', 'gender').prefetch_related(
            'variants__size', 'variants__color', 'images'  # ✅ Загружаем и размеры, и цвета
        ),
        slug=slug,
        is_active=True
    )

    # Варианты товара
    variants = product.variants.all()

    # Отзывы
    reviews = product.reviews.filter(is_moderated=True).select_related('user').order_by('-created_at')
    avg_rating = product.get_average_rating()

    # Похожие товары
    related_products = Product.objects.filter(
        category=product.category,
        is_active=True
    ).exclude(
        id=product.id
    ).select_related('category')[:4]

    # Проверка в избранном
    in_wishlist = False
    if request.user.is_authenticated:
        in_wishlist = Wishlist.objects.filter(
            user=request.user,
            product=product
        ).exists()

    # Форма отзыва
    review_form = ReviewForm()

    # Проверка может ли пользователь оставить отзыв
    can_review = False
    if request.user.is_authenticated:
        # Проверяем, есть ли у пользователя заказы этого товара (подтвержденные или доставленные)
        has_purchased = Order.objects.filter(
            user=request.user,
            items__variant__product=product,
            status__name__in=['confirmed', 'delivered', 'paid']
        ).exists()
        has_reviewed = Review.objects.filter(
            user=request.user,
            product=product
        ).exists()
        can_review = has_purchased and not has_reviewed

    context = {
        'product': product,
        'variants': variants,
        'reviews': reviews,
        'avg_rating': avg_rating,
        'related_products': related_products,
        'in_wishlist': in_wishlist,
        'review_form': review_form,
        'can_review': can_review,
    }

    return render(request, "pages/product_detail.html", context)


@login_required
def add_review(request, product_id):
    """Добавление отзыва"""
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        rating = request.POST.get('rating')
        comment = request.POST.get('comment')

        # Проверяем, может ли пользоват��ль оставить отзыв
        has_purchased = Order.objects.filter(
            user=request.user,
            items__variant__product=product,
            status__name__in=['confirmed', 'delivered', 'paid']
        ).exists()

        if not has_purchased:
            messages.error(request, 'Вы можете оставить отзыв только после оформления заказа с этим товаром')
            return redirect('product_detail', slug=product.slug)

        # Проверяем, не оставлял ли уже отзыв
        if Review.objects.filter(user=request.user, product=product).exists():
            messages.error(request, 'Вы уже оставили отзыв на этот товар')
            return redirect('product_detail', slug=product.slug)

        try:
            review = Review.objects.create(
                product=product,
                user=request.user,
                rating=rating,
                comment=comment,
                is_moderated=False
            )
            messages.success(request, 'Спасибо! Ваш отзыв отправлен на модерацию')
            logger.info(f"Review added for product {product.id} by user {request.user.id}")
        except Exception as e:
            messages.error(request, f'Ошибка при добавлении отзыва: {e}')

    return redirect('product_detail', slug=product.slug)


@login_required
def toggle_wishlist(request, product_id):
    """Добавление/удаление из избранного"""
    product = get_object_or_404(Product, id=product_id)

    wishlist_item, created = Wishlist.objects.get_or_create(
        user=request.user,
        product=product
    )

    if not created:
        wishlist_item.delete()
        message = 'Товар удален из избранного'
        in_wishlist = False
    else:
        message = 'Товар добавлен в избранное'
        in_wishlist = True

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'in_wishlist': in_wishlist,
            'message': message,
            'wishlist_count': request.user.wishlist.count()
        })

    messages.success(request, message)
    return redirect(request.META.get('HTTP_REFERER', 'catalog'))

# --- КОРЗИНА ---
@login_required
def cart_view(request):
    """Просмотр корзины"""
    cart, created = Cart.objects.get_or_create(user=request.user)
    return render(request, "pages/cart.html", {"cart": cart})


@login_required
def add_to_cart(request, variant_id):
    """Добавление товара в корзину"""
    variant = get_object_or_404(ProductVariant, id=variant_id)

    if variant.stock_quantity < 1:
        messages.error(request, 'Товара нет в наличии')
        return redirect(request.META.get('HTTP_REFERER', 'catalog'))

    cart, created = Cart.objects.get_or_create(user=request.user)

    cart_item, item_created = CartItem.objects.get_or_create(
        cart=cart,
        variant=variant,
        defaults={'quantity': 1}
    )

    if not item_created:
        if cart_item.quantity + 1 > variant.stock_quantity:
            messages.error(request, f'Доступно только {variant.stock_quantity} шт.')
        else:
            cart_item.quantity += 1
            cart_item.save()
            messages.success(request, 'Количество товара увеличено')
    else:
        messages.success(request, 'Товар добавлен в корзину')

    logger.info(f"Added to cart: user={request.user.id}, variant={variant_id}")

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({
            'success': True,
            'cart_total': str(cart.get_total_price()),
            'cart_items': cart.get_total_items(),
            'message': 'Товар добавлен в корзину'
        })

    return redirect('cart')


@login_required
def update_cart_item(request, item_id):
    """Обновление количества товара в корзине"""
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)

    if request.method == 'POST':
        try:
            quantity = int(request.POST.get('quantity', 1))
        except ValueError:
            quantity = 1

        if quantity <= 0:
            cart_item.delete()
            message = 'Товар удален из корзины'
        else:
            if quantity > cart_item.variant.stock_quantity:
                messages.error(request, f'Доступно только {cart_item.variant.stock_quantity} шт.')
                return redirect('cart')

            cart_item.quantity = quantity
            cart_item.save()
            message = 'Количество обновлено'

        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            cart = cart_item.cart
            return JsonResponse({
                'success': True,
                'message': message,
                'item_total': str(cart_item.get_total_price()),
                'cart_total': str(cart.get_total_price()),
                'cart_items': cart.get_total_items()
            })

    return redirect('cart')


@login_required
def remove_from_cart(request, item_id):
    """Удаление товара из корзины"""
    cart_item = get_object_or_404(CartItem, id=item_id, cart__user=request.user)
    cart_item.delete()

    messages.success(request, 'Товар удален из корзины')

    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        cart = Cart.objects.get(user=request.user)
        return JsonResponse({
            'success': True,
            'cart_total': str(cart.get_total_price()),
            'cart_items': cart.get_total_items()
        })

    return redirect('cart')


@login_required
def clear_cart(request):
    """Очистка корзины"""
    if request.method != 'POST':
        return redirect('cart')

    try:
        cart = Cart.objects.get(user=request.user)
        cart.items.all().delete()
        messages.success(request, 'Корзина оч��щена')
    except Cart.DoesNotExist:
        messages.info(request, 'Корзина уже пуста')

    return redirect('cart')


# --- ИЗБРАННОЕ ---
@login_required
def wishlist_view(request):
    """Просмотр избранного"""
    items = Wishlist.objects.filter(user=request.user).select_related(
        'product', 'product__category'
    ).prefetch_related(
        'product__images', 'product__variants'
    ).order_by('-added_at')

    return render(request, "pages/wishlist.html", {"wishlist_items": items})


# --- ЗАКАЗЫ ---
@login_required
def checkout(request):
    """Оформление заказа"""
    cart = get_object_or_404(Cart, user=request.user)

    if cart.get_total_items() == 0:
        messages.error(request, 'Корзина пуста')
        return redirect('cart')

    if request.method == "POST":
        form = CheckoutForm(request.POST)
        if form.is_valid():
            # Создаем заказ
            status_created, _ = OrderStatus.objects.get_or_create(name='created')

            order = Order.objects.create(
                user=request.user,
                total_amount=cart.get_total_price(),
                delivery_address=form.cleaned_data['delivery_address'],
                status=status_created,
                comment=form.cleaned_data.get('comment', '')
            )

            # Переносим товары из корзины в заказ
            for cart_item in cart.items.select_related('variant__product').all():
                order.items.create(
                    variant=cart_item.variant,
                    quantity=cart_item.quantity,
                    price_per_unit=cart_item.variant.price
                )

            logger.info(f"Order created: {order.order_number} by user {request.user.id}")

            # Получаем способ оплаты из формы
            payment_method = request.POST.get('payment_method', 'yookassa')

            # Сохраняем способ оплаты в заказ
            order.payment_method = payment_method
            order.save()

            if payment_method == 'yookassa':
                # Онлайн-оплата через ЮKassa
                # Не очищаем корзину и не уменьшаем склад - это произойдет после оплаты
                return redirect('payment', order_id=order.id)
            else:
                # Оплата наличными при получении
                # Списываем товары со склада
                for order_item in order.items.select_related('variant').all():
                    variant = order_item.variant
                    variant.stock_quantity -= order_item.quantity
                    if variant.stock_quantity < 0:
                        variant.stock_quantity = 0
                    variant.save()

                # Очищаем корзину
                cart.items.all().delete()

                # Меняем статус заказа на "подтвержден" или оставляем "created"
                # В зависимости от вашей логики
                status_confirmed, _ = OrderStatus.objects.get_or_create(name='confirmed')
                order.status = status_confirmed
                order.save()

                # Отправляем email с подтверждением заказа
                try:
                    from .utils import send_order_confirmation_email
                    send_order_confirmation_email(order, request)
                except Exception as e:
                    logger.error(f"Failed to send order confirmation email: {e}")

                messages.success(request, 'Заказ оформлен! Оплата наличными при получении.')
                return redirect('order_detail', order_id=order.id)
    else:
        initial = {}
        if request.user.orders.exists():
            last_order = request.user.orders.order_by('-created_at').first()
            initial['delivery_address'] = last_order.delivery_address

        form = CheckoutForm(initial=initial)

    return render(request, "pages/checkout.html", {
        "form": form,
        "cart": cart
    })

@login_required
def order_history(request):
    """История заказов"""
    orders = Order.objects.filter(
        user=request.user
    ).select_related(
        'status'
    ).prefetch_related(
        'items__variant__product'
    ).order_by('-created_at')

    return render(request, "pages/order_history.html", {"orders": orders})


@login_required
def order_detail(request, order_id):
    """Детали заказа"""
    order = get_object_or_404(
        Order.objects.select_related(
            'status', 'user'
        ).prefetch_related(
            'items__variant__product', 'items__variant__size',
            'transactions'
        ),
        id=order_id,
        user=request.user
    )

    return render(request, "pages/order_detail.html", {"order": order})


# --- ПЛАТЕЖИ ---
@login_required
def payment_process(request, order_id):
    """Обработка платежа через ЮKassa"""
    order = get_object_or_404(Order, id=order_id, user=request.user)

    # Проверяем, что заказ еще не оплачен
    if order.status.name in ['paid', 'shipped', 'delivered']:
        messages.warning(request, 'Заказ уже оплачен')
        return redirect('order_detail', order_id=order.id)

    # Обрабатываем платеж
    payment_url = payment_service.process_payment(order, request)

    if payment_url:
        # Перенаправляем на страницу оплаты ЮKassa
        return redirect(payment_url)
    else:
        messages.error(request, 'Не удалось создать платеж. Пожалуйста, попробуйте позже.')
        return redirect('checkout')


@login_required
def payment_result(request, order_id):
    """Результат оплаты (возврат с ЮKassa)"""
    order = get_object_or_404(Order, id=order_id, user=request.user)

    # Получаем последнюю транзакцию
    transaction = order.transactions.filter(payment_system='YooKassa').order_by('-created_at').first()

    if transaction:
        # Получаем информацию о платеже
        payment = payment_service.yookassa.get_payment_info(transaction.external_id)

        if payment:
            if payment.status == 'succeeded':
                payment_service.handle_successful_payment(transaction.external_id)
                messages.success(request, 'Оплата прошла успешно!')
            elif payment.status == 'pending':
                messages.info(request, 'Платеж обрабатывается. Мы уведомим вас о результате.')
            elif payment.status == 'canceled':
                payment_service.handle_failed_payment(transaction.external_id)
                messages.error(request, 'Платеж был отменен. Пожалуйста, попробуйте снова.')
            else:
                messages.warning(request, f'Статус платежа: {payment.status}')
        else:
            messages.error(request, 'Не удалось получить информацию о платеже')
    else:
        messages.error(request, 'Транзакция не найдена')

    return redirect('order_detail', order_id=order.id)


@csrf_exempt
def yookassa_webhook(request):
    """Webhook для получения уведомлений от ЮKassa"""
    if request.method == 'POST':
        try:
            # Получаем тело запроса
            body = request.body.decode('utf-8')

            # Обрабатываем webhook
            success = payment_service.process_webhook(body)

            if success:
                return JsonResponse({'status': 'success'})
            else:
                return JsonResponse({'status': 'error', 'message': 'Webhook processing failed'}, status=500)

        except Exception as e:
            logger.error(f"Webhook error: {e}")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'method not allowed'}, status=405)


def register_view(request):
    """Регистрация"""
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save()

            # Отправляем письмо для подтверждения email
            if send_verification_email(user, request):
                messages.success(request,
                                 'Регистрация прошла успешно! На вашу почту отправлено письмо для подтверждения email.')
            else:
                messages.warning(request,
                                 'Регистрация прошла, но не удалось отправить письмо подтверждения. Свяжитесь с поддержкой.')

            return redirect("login")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"{field}: {error}")
    else:
        form = RegisterForm()

    return render(request, "pages/register.html", {"form": form})


def verify_email(request, token):
    """Подтверждение email"""
    try:
        verification = EmailVerification.objects.get(token=token, is_used=False)

        if verification.is_valid():
            # Активируем пользователя
            user = verification.user
            user.is_active = True
            user.save()

            # Помечаем токен как использованный
            verification.is_used = True
            verification.save()

            messages.success(request, 'Email успешно подтвержден! Теперь вы можете войти в аккаунт.')
            logger.info(f"Email verified for user {user.email}")
        else:
            messages.error(request, 'Ссылка устарела. Запросите подтверждение заново.')

    except EmailVerification.DoesNotExist:
        messages.error(request, 'Неверная ссылка подтверждения.')

    return redirect('login')


def resend_verification(request):
    """По��торная отправка письма подтверждения"""
    if request.method == 'POST':
        email = request.POST.get('email')

        try:
            user = User.objects.get(email=email, is_active=False)

            # Удаляем старые неиспользованные токены
            EmailVerification.objects.filter(user=user, is_used=False).delete()

            # Отправляем новое письмо
            if send_verification_email(user, request):
                messages.success(request, 'Письмо подтверждения отправлено повторно. Проверьте вашу почту.')
            else:
                messages.error(request, 'Не удалось отправить письмо. Попробуйте позже.')

        except User.DoesNotExist:
            messages.error(request, 'Пользователь с таким email не найден или уже активирован.')

    return render(request, "pages/resend_verification.html")


def forgot_password(request):
    """Запрос на сброс пароля"""
    if request.method == 'POST':
        email = request.POST.get('email')

        try:
            user = User.objects.get(email=email, is_active=True)

            if send_password_reset_email(user, request):
                messages.success(request, 'Письмо для сброса пароля отправлено на вашу почту.')
            else:
                messages.error(request, 'Не удалось отправить письмо. Попробуйте позже.')

        except User.DoesNotExist:
            messages.error(request, 'Пользователь с таким email не найден.')

    return render(request, "pages/forgot_password.html")


def reset_password(request, token):
    """Сброс пароля"""
    try:
        verification = EmailVerification.objects.get(token=token, is_used=False)

        if not verification.is_valid():
            messages.error(request, 'Ссылка устарела. Запросите сброс пароля заново.')
            return redirect('forgot_password')

        if request.method == 'POST':
            password1 = request.POST.get('password1')
            password2 = request.POST.get('password2')

            if password1 != password2:
                messages.error(request, 'Пароли не совпадают')
            elif len(password1) < 8:
                messages.error(request, 'Пароль должен содержать минимум 8 символов')
            else:
                user = verification.user
                user.set_password(password1)
                user.save()

                verification.is_used = True
                verification.save()

                messages.success(request, 'Пароль успешно изменен! Теперь вы можете войти.')
                return redirect('login')

        return render(request, "pages/reset_password.html", {'token': token})

    except EmailVerification.DoesNotExist:
        messages.error(request, 'Неверная ссылка для сброса пароля.')
        return redirect('forgot_password')


def login_view(request):
    """Вход в систему"""
    if request.user.is_authenticated:
        return redirect('home')

    if request.method == "POST":
        form = LoginForm(request.POST)
        if form.is_valid():
            user = form.user
            login(request, user)
            messages.success(request, f'Добро пожаловать, {user.get_full_name() or user.email}!')

            next_url = request.GET.get('next', 'home')
            return redirect(next_url)
        else:
            messages.error(request, 'Неверный email или пароль')
    else:
        form = LoginForm()

    return render(request, "pages/login.html", {"form": form})


def logout_view(request):
    """Выход из системы"""
    logout(request)
    messages.info(request, 'Вы вышли из системы')
    return redirect("home")


@login_required
def profile_view(request):
    """Профиль пользователя"""
    orders = Order.objects.filter(user=request.user).select_related(
        'status'
    ).order_by('-created_at')[:5]

    return render(request, "pages/profile.html", {
        "orders": orders,
        "user": request.user
    })


@login_required
def profile_edit(request):
    """Редактирование профиля"""
    if request.method == "POST":
        form = UserProfileForm(request.POST, instance=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Профиль успешно обновлен')
            return redirect('profile')
        else:
            messages.error(request, 'Пожалуйста, исправьте ошибки в форме')
    else:
        form = UserProfileForm(instance=request.user)

    return render(request, "pages/profile_edit.html", {
        "user": request.user,
        "form": form,
    })


@login_required
def change_password(request):
    """Смена пароля"""
    if request.method == "POST":
        form = ChangePasswordForm(request.user, request.POST)
        if form.is_valid():
            form.save()
            # Повторная аутентификация, чтобы сессия не сбросилась
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, request.user)
            messages.success(request, 'Пароль успешно изменен')
            return redirect('profile')
        else:
            # Собираем ошибки из формы для отображения
            for error in form.non_field_errors():
                messages.error(request, error)
            for field, errors in form.errors.items():
                if field != '__all__':
                    for error in errors:
                        messages.error(request, error)
            return redirect('profile_edit')

    return redirect('profile_edit')


# --- МОДЕРАЦИЯ ОТЗЫВОВ ---
@user_passes_test(is_moderator)
def moderate_reviews(request):
    """Модерац��я отзывов"""
    pending_reviews = Review.objects.filter(is_moderated=False).select_related('product', 'user').order_by(
        '-created_at')
    moderated_reviews = Review.objects.filter(is_moderated=True).select_related('product', 'user',
                                                                                'moderated_by').order_by(
        '-moderated_at')[:20]

    return render(request, "pages/moderate_reviews.html", {
        'pending_reviews': pending_reviews,
        'moderated_reviews': moderated_reviews
    })


@user_passes_test(is_moderator)
def approve_review(request, review_id):
    """Одобрить отзыв"""
    review = get_object_or_404(Review, id=review_id)
    review.is_moderated = True
    review.moderated_by = request.user
    review.moderated_at = timezone.now()
    review.save()

    messages.success(request, f'Отзыв #{review_id} одобрен')
    return redirect('moderate_reviews')


@user_passes_test(is_moderator)
def reject_review(request, review_id):
    """Отклонить отзыв"""
    review = get_object_or_404(Review, id=review_id)
    review.delete()

    messages.success(request, f'Отзыв #{review_id} удален')
    return redirect('moderate_reviews')


# --- УПРАВЛЕНИЕ ТОВАРАМИ ---
@user_passes_test(is_moderator)
def manage_products(request):
    """Управление товарами"""
    products = Product.objects.all().select_related('category').prefetch_related('variants', 'images').order_by(
        '-created_at')
    categories = Category.objects.all()

    # Фильтрация
    category_id = request.GET.get('category')
    if category_id:
        products = products.filter(category_id=category_id)

    search = request.GET.get('search')
    if search:
        products = products.filter(name__icontains=search)

    return render(request, "pages/manage_products.html", {
        'products': products,
        'categories': categories
    })


@user_passes_test(is_moderator)
def edit_product(request, product_id=None):
    """Редактирование/создание товара"""
    if product_id:
        product = get_object_or_404(Product, id=product_id)
    else:
        product = None

    if request.method == 'POST':
        # Базовая логика сохранения товара
        name = request.POST.get('name')
        description = request.POST.get('description')
        price_str = request.POST.get('price')
        category_id = request.POST.get('category')
        gender_id = request.POST.get('gender')
        material = request.POST.get('material')
        care_instructions = request.POST.get('care_instructions')
        is_active = request.POST.get('is_active') == 'on'

        # Валидация цены
        try:
            price = Decimal(price_str)
            # Проверяем, что цена не превышает максимально допустимую
            max_price = Decimal('99999999.99')
            if price > max_price:
                messages.error(request, f'Цена слишком большая! Максимально допустимая цена: 99,999,999.99 ₽')
                return render(request, "pages/edit_product.html", {
                    'product': product,
                    'categories': Category.objects.all(),
                    'genders': Gender.objects.all(),
                    'sizes': Size.objects.all(),
                    'colors': Color.objects.all()
                })
            if price < 0:
                messages.error(request, 'Цена не может быть отрицательной')
                return render(request, "pages/edit_product.html", {
                    'product': product,
                    'categories': Category.objects.all(),
                    'genders': Gender.objects.all(),
                    'sizes': Size.objects.all(),
                    'colors': Color.objects.all()
                })
        except (ValueError, DecimalException):
            messages.error(request, 'Введите корректное числовое значение цены')
            return render(request, "pages/edit_product.html", {
                'product': product,
                'categories': Category.objects.all(),
                'genders': Gender.objects.all(),
                'sizes': Size.objects.all(),
                'colors': Color.objects.all()
            })

        if product:
            # Обновление существующего товара
            product.name = name
            product.description = description
            product.price = price
            if category_id:
                product.category_id = category_id
            if gender_id:
                product.gender_id = gender_id
            product.material = material
            product.care_instructions = care_instructions
            product.is_active = is_active
            product.save()

            # Обработка вариантов товара
            sizes = request.POST.getlist('size[]')
            colors = request.POST.getlist('color[]')
            prices = request.POST.getlist('price[]')
            stocks = request.POST.getlist('stock[]')

            # Получаем существующие варианты
            existing_variants = {f"{v.size_id}_{v.color_id}": v for v in product.variants.all()}
            processed_keys = set()

            # Обновляем или создаем новые варианты
            for i in range(len(sizes)):
                if sizes[i] and prices[i] and stocks[i]:
                    try:
                        variant_price = Decimal(prices[i])
                        # Проверяем цену варианта
                        if variant_price > Decimal('99999999.99'):
                            messages.warning(request, f'Цена варианта {i+1} слишком большая, установлена максимальная')
                            variant_price = Decimal('99999999.99')
                        if variant_price < 0:
                            messages.warning(request, f'Цена варианта {i+1} отрицательная, установлена 0')
                            variant_price = Decimal('0')
                    except (ValueError, DecimalException):
                        messages.error(request, f'Введите корректную цену для варианта {i+1}')
                        continue

                    size_id = sizes[i] if sizes[i] else None
                    color_id = colors[i] if colors[i] else None
                    variant_key = f"{size_id}_{color_id}"

                    if variant_key in existing_variants:
                        # Обновляем существующий вариант
                        variant = existing_variants[variant_key]
                        variant.price = variant_price
                        variant.stock_quantity = int(stocks[i]) if stocks[i] else 0
                        variant.save()
                        processed_keys.add(variant_key)
                    else:
                        # Создаем новый вариант
                        ProductVariant.objects.create(
                            product=product,
                            size_id=size_id,
                            color_id=color_id,
                            price=variant_price,
                            stock_quantity=int(stocks[i]) if stocks[i] else 0
                        )

            messages.success(request, 'Товар успешно обновлен')
        else:
            # Создание нового товара
            from django.utils.text import slugify
            import random

            # Создаем уникальный slug
            base_slug = slugify(name)
            slug = base_slug
            while Product.objects.filter(slug=slug).exists():
                slug = f"{base_slug}-{random.randint(1, 9999)}"

            product = Product.objects.create(
                name=name,
                slug=slug,
                description=description,
                price=price,
                category_id=category_id if category_id else None,
                gender_id=gender_id if gender_id else None,
                material=material,
                care_instructions=care_instructions,
                is_active=is_active
            )

            # Обработка вариантов товара для нового товара
            sizes = request.POST.getlist('size[]')
            colors = request.POST.getlist('color[]')
            prices = request.POST.getlist('price[]')
            stocks = request.POST.getlist('stock[]')

            for i in range(len(sizes)):
                if sizes[i] and prices[i] and stocks[i]:
                    try:
                        variant_price = Decimal(prices[i])
                        # Проверяем цену варианта
                        if variant_price > Decimal('99999999.99'):
                            messages.warning(request, f'Цена варианта {i+1} слишком большая, установлена максимальная')
                            variant_price = Decimal('99999999.99')
                        if variant_price < 0:
                            messages.warning(request, f'Цена варианта {i+1} отрицательная, установлена 0')
                            variant_price = Decimal('0')
                    except (ValueError, DecimalException):
                        messages.error(request, f'Введите корректную цену для варианта {i+1}')
                        continue

                    ProductVariant.objects.create(
                        product=product,
                        size_id=sizes[i] if sizes[i] else None,
                        color_id=colors[i] if colors[i] else None,
                        price=variant_price,
                        stock_quantity=int(stocks[i]) if stocks[i] else 0
                    )

            messages.success(request, 'Товар успешно создан')

        # Обработка изображений
        if request.FILES.getlist('images'):
            for image in request.FILES.getlist('images'):
                ProductImage.objects.create(
                    product=product,
                    image=image,
                    is_main=not product.images.exists()
                )

        return redirect('manage_products')

    return render(request, "pages/edit_product.html", {
        'product': product,
        'categories': Category.objects.all(),
        'genders': Gender.objects.all(),
        'sizes': Size.objects.all(),
        'colors': Color.objects.all()
    })


# --- УПРАВЛЕНИЕ ЗАКАЗАМИ ---
@user_passes_test(is_moderator)
def manage_orders(request):
    """Управление заказами"""
    status_filter = request.GET.get('status')
    orders = Order.objects.all().select_related('user', 'status').prefetch_related('items').order_by('-created_at')

    if status_filter:
        orders = orders.filter(status__name=status_filter)

    statuses = OrderStatus.objects.all()

    return render(request, "pages/manage_orders.html", {
        'orders': orders,
        'statuses': statuses,
        'current_status': status_filter
    })


@user_passes_test(is_moderator)
def update_order_status(request, order_id):
    """Обновление статуса заказа"""
    order = get_object_or_404(Order, id=order_id)

    if request.method == 'POST':
        status_name = request.POST.get('status')
        new_status = get_object_or_404(OrderStatus, name=status_name)
        order.status = new_status
        order.save()

        messages.success(request, f'Статус заказа #{order.order_number} обновлен')

    return redirect('manage_orders')


@user_passes_test(is_moderator)
def delete_product(request, product_id):
    """Удаление товара (только если нет связанных заказов)"""
    product = get_object_or_404(Product, id=product_id)

    # Проверяем, есть ли заказы с этим товаром
    has_orders = OrderItem.objects.filter(variant__product=product).exists()

    if has_orders:
        # Если есть заказы, просто деактивируем товар
        product.is_active = False
        product.save()
        messages.warning(request, f'Товар "{product.name}" деактивирован, так как есть связанные заказы')
    else:
        # Если заказов нет, удаляем полностью
        try:
            # Сначала удаляем связанные изображения из файловой системы
            for image in product.images.all():
                if image.image:
                    image.image.delete()  # Удаляем файл
            # Удаляем варианты товара
            product.variants.all().delete()
            # Удаляем сам товар
            product.delete()
            messages.success(request, f'Товар "{product.name}" полностью удален')
        except Exception as e:
            messages.error(request, f'Ошибка при удалении товара: {str(e)}')

    return redirect('manage_products')


# --- АДМИН ПАНЕЛЬ ---
@user_passes_test(is_admin)
def admin_dashboard(request):
    """Панель администратора с аналитикой"""
    try:
        # Общая статистика
        total_users = User.objects.count()
        total_products = Product.objects.count()
        total_orders = Order.objects.count()

        # Выручка только по оплаченным заказам
        paid_orders = Order.objects.filter(status__name='paid')
        total_revenue = paid_orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0

        # Статистика за последние 30 дней
        last_month = timezone.now() - timedelta(days=30)
        recent_orders = Order.objects.filter(created_at__gte=last_month).count()
        recent_revenue = Order.objects.filter(
            created_at__gte=last_month,
            status__name='paid'
        ).aggregate(Sum('total_amount'))['total_amount__sum'] or 0

        # Топ товаров
        top_products = Product.objects.annotate(
            order_count=Count('variants__orderitem')
        ).filter(order_count__gt=0).order_by('-order_count')[:5]

        # Статусы заказов
        order_statuses = OrderStatus.objects.annotate(
            count=Count('orders')
        )

        context = {
            'total_users': total_users,
            'total_products': total_products,
            'total_orders': total_orders,
            'total_revenue': total_revenue,
            'recent_orders': recent_orders,
            'recent_revenue': recent_revenue,
            'top_products': top_products,
            'order_statuses': order_statuses,
        }

        return render(request, "pages/admin_dashboard.html", context)

    except Exception as e:
        # Логируем ошибку
        logger.error(f"Dashboard error: {e}")
        messages.error(request, f"Ошибка загрузки панели управления: {e}")
        return redirect('home')


@user_passes_test(is_admin)
def export_orders(request):
    """Экспорт заказов в CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="orders.csv"'

    writer = csv.writer(response)
    writer.writerow(['Номер заказа', 'Пользователь', 'Email', 'Дата', 'Сумма', 'Статус', 'Адрес'])

    orders = Order.objects.select_related('user', 'status').all()
    for order in orders:
        writer.writerow([
            order.order_number,
            order.user.get_full_name(),
            order.user.email,
            order.created_at.strftime('%d.%m.%Y %H:%M'),
            order.total_amount,
            order.status.get_name_display(),
            order.delivery_address
        ])

    return response


@user_passes_test(is_admin)
def export_orders_utf8(request):
    """Экспорт заказов в CSV с поддержкой UTF-8 BOM (для Excel)"""
    import codecs

    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="orders_utf8.csv"'

    # Добавляем BOM для правильного отображения в Excel
    response.write(codecs.BOM_UTF8)

    writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Заголовки на русском
    writer.writerow(['Номер заказа', 'Пользователь', 'Email', 'Дата', 'Сумма', 'Статус', 'Адрес'])

    orders = Order.objects.select_related('user', 'status').all()
    for order in orders:
        writer.writerow([
            order.order_number,
            order.user.get_full_name() or order.user.email,
            order.user.email,
            order.created_at.strftime('%d.%m.%Y %H:%M'),
            str(order.total_amount).replace('.', ','),  # Заменяем точку на запятую для Excel
            order.status.get_name_display(),
            order.delivery_address
        ])

    return response


@user_passes_test(is_admin)
def export_orders_excel(request):
    """Экспорт заказов в Excel с поддержкой русского текста"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        messages.error(request, 'Библиотека openpyxl не установлена. Установите: pip install openpyxl')
        return redirect('admin_dashboard')

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Заказы"

    # Заголовки
    headers = ['Номер заказа', 'Пользователь', 'Email', 'Дата', 'Сумма', 'Статус', 'Адрес доставки']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=12)
        cell.fill = PatternFill(start_color="D4A373", end_color="D4A373", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')

    # Данные
    orders = Order.objects.select_related('user', 'status').all()
    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.order_number)
        ws.cell(row=row, column=2, value=order.user.get_full_name() or order.user.email)
        ws.cell(row=row, column=3, value=order.user.email)
        ws.cell(row=row, column=4, value=order.created_at.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=5, value=float(order.total_amount))
        ws.cell(row=row, column=6, value=order.status.get_name_display())
        ws.cell(row=row, column=7, value=order.delivery_address)

        # Выравнивание
        for col in range(1, 8):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='left')

    # Автоматическая ширина колонок
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="orders.xlsx"'

    wb.save(response)
    return response


@user_passes_test(is_admin)
def export_orders_csv_windows(request):
    """Экспорт заказов в CSV для Windows (кодировка windows-1251)"""
    response = HttpResponse(content_type='text/csv; charset=windows-1251')
    response['Content-Disposition'] = 'attachment; filename="orders_win.csv"'

    writer = csv.writer(response, delimiter=';', quoting=csv.QUOTE_MINIMAL)

    # Заголовки
    writer.writerow(['Номер заказа', 'Пользователь', 'Email', 'Дата', 'Сумма', 'Статус', 'Адрес'])

    orders = Order.objects.select_related('user', 'status').all()
    for order in orders:
        writer.writerow([
            order.order_number,
            order.user.get_full_name() or order.user.email,
            order.user.email,
            order.created_at.strftime('%d.%m.%Y %H:%M'),
            str(order.total_amount).replace('.', ','),
            order.status.get_name_display(),
            order.delivery_address
        ])

    return response